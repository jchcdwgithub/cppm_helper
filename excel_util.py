from openpyxl.worksheet.table import Table
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import math
import pandas

def write_to_excel(title, data, workbook, start):
    
    current_ws = workbook.create_sheet(title=title)
    title_row_color = 'C0504D'

    for row in data:
        current_ws.append(row)

    colors = ['E6B8B7', 'F2DCDB']
    color_index = 0
    rows = current_ws.rows
    for row in rows:
        if color_index == 0:
            for cell in row:
                cell.fill = PatternFill('solid', fgColor = title_row_color)
                cell.font = Font(color='FFFFFF')
        else:
            for cell in row:
                cell.fill = PatternFill('solid', fgColor=colors[color_index%2])
                cell.font = Font(color='000000')
        color_index += 1
    
    column_widths = []
    for row in data:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]
    
    for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
        current_ws.column_dimensions[get_column_letter(i)].width = column_width + 5

    table = Table(displayName='SSID',ref=f'A{start}:F{start + len(data)}')
    current_ws.add_table(table)

def new_add_table_to_worksheet(data,worksheet,appliance_name='',table_name='',start=1):
    
    end_cell_number = len(data)
    end_cell = (end_cell_number, start + len(data[0]))
    start_cell = (1 if appliance_name == '' else 2,start)
    num_columns = len(data[0])

    current = 1 
    data_index = 0
    if appliance_name != '':
        appliance_name_cell = worksheet.cell(row=1,column=start)
        appliance_name_cell.value = appliance_name
        current += 1
    while current <= end_cell_number:
        current_column = start
        current_row_cells = []
        while current_column < start + num_columns:
            current_row_cells.append(worksheet.cell(row=current, column=current_column))
            current_column += 1
        current_data_row = data[data_index]
        for cell,data_value in zip(current_row_cells,current_data_row):
            cell.value = data_value
        current += 1
        data_index += 1 
    add_color_scheme(worksheet,start_cell,end_cell)
    column_widths = get_widest_column_widths(start-1,data)
    adjust_column_widths(worksheet,column_widths)

    return

def adjust_column_widths(worksheet,column_widths):
    ''' Adjust the column widths of the table in the worksheet based on the dictionary passed. '''

    for column_letter in column_widths:
        worksheet.column_dimensions[column_letter].width = column_widths[column_letter] + 5

def add_color_scheme(worksheet,start_cell,end_cell):
    ''' Add a red table color scheme to the cells in the worksheet. '''
    title_row_color = 'C0504D'
    colors = ['E6B8B7', 'F2DCDB']
    color_index = 0
    rows = []
    start_row,start_column = start_cell
    end_row,end_column = end_cell
    current_row = start_row 
    while current_row <= end_row:
        current_column = start_column 
        current_row_list = []
        while current_column < end_column:
            current_row_list.append(worksheet.cell(row=current_row,column=current_column))
            current_column += 1
        rows.append(current_row_list)
        current_row += 1
    for row in rows:
       if color_index == 0:
          for cell in row:
               cell.fill = PatternFill('solid', fgColor = title_row_color)
               cell.font = Font(color='FFFFFF')
       else:
            for cell in row:
               cell.fill = PatternFill('solid', fgColor=colors[color_index%2])
               cell.font = Font(color='000000')
       color_index += 1

def swap_rows_and_columns(data):
    ''' Given an array of arrays, swap the rows and columns data. '''

    num_rows = len(data)
    swapped_matrix = []
    num_cols = len(data[0])
    current_col = 0
    while current_col < num_cols:
        current_row = 0
        current_col_array = []
        while current_row < num_rows:
            current_col_array.append(data[current_row][current_col])
            current_row += 1
        current_col += 1
        swapped_matrix.append(current_col_array)

    return swapped_matrix

def calculate_column_letters(column_index):
    ''' Given a column index, return the corresponding excel column letter. '''

    column_letters = [letter.upper() for letter in 'abcdefghijklmnopqrstuvwxyz']

    if column_index < 26:
        return column_letters[column_index]
    else:
        first_letter_index = math.floor(column_index/26) - 1
        last_letter_index = column_index%26 - 1

        first_letter = column_letters[first_letter_index]
        last_letter = column_letters[last_letter_index]

        column = first_letter + last_letter
        return column

def get_widest_column_widths(data):
    ''' Returns a list of widest length of data found in each column. '''
    if len(data) == 0:
        return []
    col_widths = [0 for _ in data[0]]
    for current_list in data:
        for col,ele in enumerate(current_list):
            ele_len = len(str(ele))
            if ele_len > col_widths[col]:
                col_widths[col] = ele_len
    return [col_width+2 for col_width in col_widths]

def format_cells_in_table(color_scheme:dict[str,str], writer:pandas.ExcelWriter, worksheet, 
                          start_cell:tuple, table_name:str, table_headers:list[str], table_data:list[list]) -> None:
    ''' Takes a color scheme, applies it to the table_data table and writes the data to the excel sheet '''
    start_row, start_col = start_cell
    workbook = writer.book
    table_header_format = workbook.add_format({'bg_color':color_scheme['header_bg_color'], 'font_color':color_scheme['header_font_color']})
    colored_row_format = workbook.add_format({'bg_color':color_scheme['row_bg_color'], 'font_color':color_scheme['row_font_color']})
    worksheet
    
    for col,header in enumerate(table_headers):
        worksheet.write(start_row, start_col+col, header, table_header_format)
    
    start_row += 1
    for row_index,row in enumerate(table_data):
        if row_index%2 == 0:
            for col_index, ele in enumerate(row):
                worksheet.write(start_row+row_index, start_col+col_index, ele, colored_row_format)
        else:
            for col_index, ele in enumerate(row):
                worksheet.write(start_row+row_index, start_col+col_index, ele)

def make_headers_object(headers:list[str]) -> list[dict[str, str]]:
    '''
    Return a list of {'header': 'table header name'} 
    '''
    return [{'header':header} for header in headers]

def add_headers_row_to_data(table:dict) -> list[list]:
    '''
    Add the headers row to the table data as the first row. Returns the new list of lists.
    '''
    with_headers = [table['headers']]
    for row in table['data']:
        with_headers.append(row)
    return with_headers

def find_max_number_of_columns_among_all_tables_in_set(tables:list[dict]) -> int:
    '''
    Given a list of tables, return the table with the most columns
    '''
    max_cols = 0
    for table in tables:
        current_number_cols = len(table['headers'])
        if current_number_cols > max_cols:
            max_cols = current_number_cols
    return max_cols

def find_largest_width_among_multiple_tables(max_cols:int, tables:list[dict]) -> list[int]:
    '''
    Create a list of widest widths for all columns across multiple tables.
    '''
    widths = [0 for _ in range(max_cols)]
    for table in tables:
        table_widths = get_widest_column_widths(add_headers_row_to_data(table))
        for index,col_width in enumerate(table_widths):
            if col_width > widths[index]:
                widths[index] = col_width

    for width_index,width in enumerate(widths):
        widths[width_index] = width + 5
    return widths

def write_table_to_excel_worksheet(worksheet, table:dict[str, list[list]], start_cell:tuple, title_format) -> int:
    '''
    Writes a single table to an excel worksheet. Returns the starting row and column for the next table. 
    Worksheet is an excel worksheet object taken from one of the sheets of the writer.
    The table is a dictionary that has two keys: name and data.
        name is a string representing the table name.
        data is a list of rows with the table data. Headers are included.
    '''
    current_row, current_col= start_cell
    worksheet.write(current_row, current_col, table['name'], title_format)
    current_row += 1
    table_headers = table['headers']
    table_rows = len(table['data'])
    table_columns = make_headers_object(table_headers)
    table_style = 'Table Style Medium 2'
    if 'style' in table:
        table_style = table['style']

    worksheet.add_table(
        current_row,
        current_col,
        current_row+len(table['data']),
        current_col+len(table_headers)-1,
        {
            'columns' : table_columns,
            'data' : table['data'],
            'style' : table_style
        }
    )
    
    return current_row+table_rows+1

def write_tables_to_excel_worksheets(excel_filename: str, worksheet_names:list[str], tables_sets:list[list[list[list]]]):
    ''' Takes a set of tables and writes them onto the same sheet in an excel file.
        Each element in the Tables list corresponds to a list of tables to be written vertically
        in relation to each other. Each table is a dictionary that contains the name, headers and
        data in the table.
    '''
    with pandas.ExcelWriter(excel_filename) as writer:
        workbook = writer.book
        for worksheet_name,tables in zip(worksheet_names,tables_sets):
            worksheet = workbook.add_worksheet(worksheet_name)
            writer.sheets[worksheet_name] = worksheet
            table_title_cell_format = workbook.add_format({'bg_color': '#D9D9D9', 'font_size': 16, 'bold':True})
            next_table_row = 0
            next_table_col = 0

            for tables_in_set in tables:
                max_columns = find_max_number_of_columns_among_all_tables_in_set(tables_in_set)
                widths = find_largest_width_among_multiple_tables(max_columns, tables_in_set)
                for col_offset,width in enumerate(widths):
                    column = next_table_col + col_offset
                    worksheet.set_column(column,column,width)

                for table in tables_in_set:
                    next_table_row = write_table_to_excel_worksheet(worksheet, table, (next_table_row, next_table_col), table_title_cell_format)

                next_table_row = 0
                next_table_col += max_columns + 2
        
        writer.save()